from openpyxl import load_workbook

# Load workbook
file_path = "../content/Phishing_file_content/Phishing_Attempt.xlsx"
wb = load_workbook(file_path, data_only=True)  # Load without formulas
ws = wb.active  # Select the first sheet

# Unhide all hidden columns
for col in ws.column_dimensions:
    if ws.column_dimensions[col].hidden:
        ws.column_dimensions[col].hidden = False

# Print username:password
usernames,passwords,credentials = [],[],[]
for i,row in enumerate(ws.iter_rows(values_only=True)):
    if i == 0: # skip the first row which is the column names
        continue
    if row[-1] == None: # skip null rows
        continue
    print(str(row[-1])+':'+str(row[-2]))  
    usernames.append(str(row[-1]))
    passwords.append(str(row[-2]))
    credentials.append(str(row[-1]) + ':' + str(row[-2]))


with open('phishing_users.txt', 'w') as f:
    for user in usernames:
        f.write(user + "\n")
with open('phishing_passwords.txt', 'w') as f:
    for password in passwords:
        f.write(password + "\n")
with open('phishing_credentials.txt', 'w') as f:
    for credential in credentials:
        f.write(credential + "\n")
print("[+] Done")
