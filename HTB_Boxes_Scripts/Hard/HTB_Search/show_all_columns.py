from openpyxl import load_workbook

# Load workbook
file_path = "../content/Phishing_file_content/Phishing_Attempt.xlsx"
wb = load_workbook(file_path, data_only=True)  # Load without formulas
ws = wb.active  # Select the first sheet

# Unhide all hidden columns
for col in ws.column_dimensions:
    if ws.column_dimensions[col].hidden:
        ws.column_dimensions[col].hidden = False
        print(f"Unhiding column: {col}")

# Print all data
for row in ws.iter_rows(values_only=True):
    print(row)  # Print entire row data

wb.save("unhidden.xlsx")  # Save the unhidden version
